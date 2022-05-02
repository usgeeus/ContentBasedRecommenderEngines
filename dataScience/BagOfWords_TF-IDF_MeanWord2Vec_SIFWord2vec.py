#!/usr/bin/env python
# coding: utf-8

# In[18]:


import pandas as pd
import numpy as np


# In[19]:


data = pd.read_csv('C:/Users/Gee/Documents/카카오톡 받은 파일/books8.csv', low_memory=False)


# In[20]:


data = data.dropna(how = 'any') 
print(data.isnull().values.any())


# In[22]:


data['words'] = data[['bookName','description']].apply(lambda x: ' '.join(x), axis=1)


# In[23]:


data = data.drop(columns=['Unnamed: 0'])
data['ID'] = list(range(data.shape[0]))


# In[24]:


data['words'] = data['words'].str.replace("[^ㄱ-ㅎㅏ-ㅣ가-힣 ]","")


# In[25]:


stop = ['들','좀','과','은','도','걍','를','으로','자','에','잘','의','가','이','와','한','하다','는']


# In[26]:


data


# In[27]:


from konlpy.tag import Okt


# In[28]:


okt = Okt()
t_d = []
for i in data['words']:
    provisional = okt.morphs(i, stem=True) 
    provisional = [word for word in provisional if not word in stop] 
    t_d.append(provisional)


# In[29]:


from gensim.models import Word2Vec
w2v_model = Word2Vec(sentences = t_d, size = 300, window =5, sg = 1)


# In[15]:


data.to_excel('C:/Users/Gee/Documents/카카오톡 받은 파일/datafinal22.xlsx')


# In[16]:


data = pd.read_excel('C:/Users/Gee/Documents/카카오톡 받은 파일/dataexcercise.xlsx')


# In[30]:


def get_ISBN(index):
    return data[data.ID ==index]["ISBN"].values[0]

def recommend(data_id, sim_mat):
    r_list = list()
    likeness =  list(enumerate(sim_mat[data_id]))
    arranged_likeness = sorted(likeness,key=lambda x:x[1],reverse=True)
    
    for i in range(20): 
        ISBN = get_ISBN(arranged_likeness[i][0])
        r_list.append(ISBN)
            
    return r_list


# In[31]:


from sklearn.metrics.pairwise import cosine_similarity


# In[32]:


from sklearn.feature_extraction.text import CountVectorizer
Bag_of_words = CountVectorizer()
Bag_of_words_mat = Bag_of_words.fit_transform(data["words"])


# In[33]:


import pickle


# In[21]:


with open('C:/Users/Gee/Documents/카카오톡 받은 파일/Bag_of_words_matexcercise.pkl', 'wb') as f:
    pickle.dump(Bag_of_words_mat, f)


# In[22]:


with open('C:/Users/Gee/Documents/카카오톡 받은 파일/Bag_of_words_matexcercise.pkl', 'rb') as f:
    Bag_of_words_mat = pickle.load(f)


# In[34]:


Bag_of_words_cosine_sim = cosine_similarity(Bag_of_words_mat)


# In[35]:


import openpyxl


# In[25]:


wb0 = openpyxl.Workbook()
sheet = wb0.active
for i in range(10): 
    sheet.append(recommend(i, Bag_of_words_cosine_sim))

wb0.save("C:/Users/Gee/Documents/카카오톡 받은 파일/Bag_of_wordsexcercise.xlsx")


# In[36]:


for i in range(10):
    print(recommend(i, Bag_of_words_cosine_sim))
    print()


# In[37]:


from sklearn.feature_extraction.text import TfidfVectorizer
tf_idf = TfidfVectorizer()
tf_idf_mat = tf_idf.fit_transform(data["words"])


# In[28]:


with open('C:/Users/Gee/Documents/카카오톡 받은 파일/tf_idf_matexcercise.pkl', 'wb') as f:
    pickle.dump(tf_idf_mat, f)


# In[29]:


with open('C:/Users/Gee/Documents/카카오톡 받은 파일/tf_idf_matexcercise.pkl', 'rb') as f:
    tf_idf_mat = pickle.load(f)


# In[38]:


tf_idf_cosine_sim = cosine_similarity(tf_idf_mat)


# In[40]:


wb1 = openpyxl.Workbook()
sheet = wb1.active
for i in range(10):
    sheet.append(recommend(i, tf_idf_cosine_sim))

wb1.save("C:/Users/Gee/Documents/카카오톡 받은 파일/tf_idfexcercjse10.xlsx")


# In[41]:


for i in range(10):
    print(recommend(i, tf_idf_cosine_sim))
    print()


# In[42]:


from gensim.models import Word2Vec
from sklearn.decomposition import PCA
from matplotlib import pyplot


# In[43]:


class MyTokenizer:
    def __init__(self):
        pass
    
    def fit(self, X, y=None):
        return self
    
    def transform(self, X):
        transformed_X = []
        for document in X:
            tokenized_doc = []
            for sent in nltk.sent_tokenize(document):
                tokenized_doc += nltk.word_tokenize(sent)
            transformed_X.append(np.array(tokenized_doc))
        return np.array(transformed_X)
    
    def fit_transform(self, X, y=None):
        return self.transform(X)

class AverageEmbedVect(object):
    def __init__(self, word2vec):
        self.word2vec = word2vec
        # if a text is empty we should return a vector of zeros
        # with the same dimensionality as all the other vectors
        self.dim = len(word2vec.wv.syn0[0])

    def fit(self, X, y=None):
        return self

    def transform(self, X):
        X = MyTokenizer().fit_transform(X)
        
        return np.array([
            np.mean([self.word2vec.wv[w] for w in words if w in self.word2vec.wv]
                    or [np.zeros(self.dim)], axis=0)
            for words in X
        ])
    
    def fit_transform(self, X, y=None):
        return self.transform(X)


# In[44]:


words_list = list(data.words)


# In[45]:


import nltk
nltk.download('punkt')


# In[46]:


Average_embedding_vectorizer = AverageEmbedVect(w2v_model)
Average_embedded = Average_embedding_vectorizer.fit_transform(data['words'])


# In[40]:


np.savetxt('C:/Users/Gee/Documents/카카오톡 받은 파일/Average_embeddedexcercise.txt', Average_embedded)


# In[41]:


AverageEmbedVect = np.loadtxt('C:/Users/Gee/Documents/카카오톡 받은 파일/Average_embeddedexcercise.txt')


# In[47]:


Mean_w2v_cosine_sim = cosine_similarity(Average_embedded)


# In[43]:


wb2 = openpyxl.Workbook()
sheet = wb2.active
for i in range(10):
    sheet.append(recommend(i, Mean_w2v_cosine_sim))

wb2.save("C:/Users/Gee/Documents/카카오톡 받은 파일/mw2vexcercise.xlsx")


# In[48]:


for i in range(10):
    print(recommend(i, Mean_w2v_cosine_sim))
    print()


# In[49]:


from scipy.sparse.linalg import svds
from sklearn.decomposition import TruncatedSVD


# In[50]:


def rem_fir_princ_comp(X):
    svd = TruncatedSVD(n_components=1, n_iter=7, random_state=0)
    svd.fit(X)
    pc = svd.components_
    XX = X - X.dot(pc.transpose()) * pc
    return XX

def sif(sent, a=0.001, word2vec_model=w2v_model):
    word_counter = {}
    lines = []
    total_count = 0
    no_of_lines = 0
    
    for s in sent:
        for w in s:
            if w in word_counter:
                word_counter[w] = word_counter[w] + 1
            else:
                word_counter[w] = 1
        total_count = total_count + len(s)
        no_of_lines = no_of_lines + 1
    
    sents_emd = []
    for s in sent:
        sent_emd = []
        for word in s:
            if word in word2vec_model:
                emd = (a/(a + (word_counter[word]/total_count)))*word2vec_model[word]
                sent_emd.append(emd)
        sum_ = np.array(sent_emd).sum(axis=0)
        line_emd = sum_/float(no_of_lines)
        sents_emd.append(line_emd)
        
    new_sents_emb = rem_fir_princ_comp(np.array(sents_emd, dtype = object))
    return new_sents_emb


# In[51]:


sif_words_emd = sif(words_list)


# In[48]:


np.savetxt('C:/Users/Gee/Documents/카카오톡 받은 파일/sif_words_emdfinal22.txt', sif_words_emd)


# In[49]:


sif_words_emd = np.loadtxt('C:/Users/Gee/Documents/카카오톡 받은 파일/sif_words_emdexcercise.txt')


# In[52]:


sif_cosine_sim = cosine_similarity(sif_words_emd)


# In[51]:


wb3 = openpyxl.Workbook()
sheet = wb3.active
for i in range(10):
    sheet.append(recommend(i, sif_cosine_sim))

wb3.save("C:/Users/Gee/Documents/카카오톡 받은 파일/sifexcercise.xlsx")


# In[53]:


for i in range(10):
    print(recommend(i, sif_cosine_sim))
    print()


# In[54]:


from gensim.models import KeyedVectors
w2v_model.wv.save_word2vec_format('w2v_model2') 
loaded_model = KeyedVectors.load_word2vec_format("w2v_model2")


# In[ ]:


model_result = loaded_model.most_similar("남자")


# In[56]:


get_ipython().system('python -m gensim.scripts.word2vec2tensor --input w2v_model2 --output w2v_model2')


# In[ ]:




